import calendar
import logging
import os
from datetime import date, datetime, timedelta
from datetime import datetime

from flask import (Blueprint, jsonify, render_template, request, send_file,
                   session)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_

from config import Config
from models import Employee, WorkLog, db
from routes.auth import role_required

logger = logging.getLogger(__name__)

worklog_bp = Blueprint("worklog", __name__)


def calculate_employee_summary(employee_id):
    """Calculate summary of work hours, days, and holidays for an employee."""
    work_logs = WorkLog.query.filter_by(employee_id=employee_id).all()
    total_hours = sum(log.worked_hours or 0 for log in work_logs)
    working_days = sum(
        1 for log in work_logs if log.worked_hours and log.worked_hours > 0
    )
    paid_holidays = sum(1 for log in work_logs if log.holidays == "paid")
    unpaid_holidays = sum(1 for log in work_logs if log.holidays == "unpaid")
    total_hours = round(total_hours)
    overtime_hours = max(0, total_hours - (working_days * 8))

    Config.logger.info(f" Summary for employee_id={employee_id}:")
    Config.logger.info(
        f" total_hours={total_hours}, working_days={working_days}, overtime_hours={overtime_hours}"
    )
    Config.logger.info(
        f" paid_holidays={paid_holidays}, unpaid_holidays={unpaid_holidays}"
    )

    return {
        "total_hours": total_hours,
        "working_days": working_days,
        "overtime_hours": overtime_hours,
        "paid_holidays": paid_holidays,
        "unpaid_holidays": unpaid_holidays,
    }


@worklog_bp.route("/work", methods=["GET"])
@role_required(["admin"])
def work():
    try:
        requested = request.args.get("filter")
        allowed = {
            "today",
            "yesterday",
            "last7days",
            "last30days",
            "thismonth",
            "lastmonth",
        }

        if requested in allowed:
            filter_type = requested
            session["filter_type"] = filter_type
        elif requested == "personalizado":
            filter_type = "personalizado"
            session.pop("filter_type", None)
        else:
            prev = session.get("filter_type")
            filter_type = prev if prev in allowed else "thismonth"
            session["filter_type"] = filter_type

        today = date.today()
        start_date = end_date = None

        if filter_type == "today":
            start_date = end_date = today
        elif filter_type == "yesterday":
            start_date = end_date = today - timedelta(days=1)
        elif filter_type == "last7days":
            start_date = today - timedelta(days=6)
            end_date = today
        elif filter_type == "last30days":
            start_date = today - timedelta(days=29)
            end_date = today
        elif filter_type == "thismonth":
            start_date = today.replace(day=1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = today.replace(day=last_day)
        elif filter_type == "lastmonth":
            first_day_this = today.replace(day=1)
            last_day_prev = first_day_this - timedelta(days=1)
            start_date = last_day_prev.replace(day=1)
            end_date = last_day_prev
        elif filter_type == "personalizado":
            start_str = request.args.get("start_date")
            end_str = request.args.get("end_date")
            if not start_str or not end_str:
                return (
                    jsonify(
                        {
                            "error": "El filtro 'personalizado' requiere 'start_date' y 'end_date'"
                        }
                    ),
                    400,
                )
            try:
                start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
            except ValueError:
                return jsonify({"error": "Formato de fecha no válido"}), 400
            if start_date > end_date:
                return (
                    jsonify(
                        {
                            "error": "La fecha de inicio no puede ser después de la fecha de finalización"
                        }
                    ),
                    400,
                )
        else:
            filter_type = "thismonth"
            start_date = today.replace(day=1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = today.replace(day=last_day)
            session["filter_type"] = filter_type

        if start_date is None or end_date is None:
            return jsonify({"error": "Faltan las fechas de filtro"}), 400

        employees = Employee.query.filter(
            and_(
                or_(Employee.end_date is None, Employee.end_date >= start_date),
                Employee.start_date <= end_date,
            )
        )

        employee_logs = []
        for emp in employees:
            logs = WorkLog.query.filter(
                and_(
                    WorkLog.employee_id == emp.id,
                    WorkLog.log_date >= start_date,
                    WorkLog.log_date <= end_date,
                )
            ).all()

            work_logs_data = []
            total_hours = working_days = paid_holidays = unpaid_holidays = (
                overtime_hours
            ) = 0

            for log in logs:
                log.calculate_worked_hours()
                wd = log.log_date
                if isinstance(wd, str):
                    wd = datetime.strptime(wd, "%Y-%m-%d").date()

                hours = float(log.worked_hours or 0)
                total_hours += hours

                if log.holidays == "workingday" and hours > 0:
                    working_days += 1
                    if hours > 8:
                        overtime_hours += hours - 8
                elif log.holidays == "paid":
                    paid_holidays += 1
                elif log.holidays == "unpaid":
                    unpaid_holidays += 1

                work_logs_data.append(
                    {
                        "id": log.id,
                        "log_date": wd.strftime("%Y-%m-%d"),
                        "check_in_time": (
                            log.check_in_time.strftime("%H:%M")
                            if log.check_in_time
                            else "--:--"
                        ),
                        "check_out_time": (
                            log.check_out_time.strftime("%H:%M")
                            if log.check_out_time
                            else "--:--"
                        ),
                        "worked_hours": round(hours, 2),
                        "holidays": log.holidays or "workingday",
                    }
                )

            employee_logs.append(
                {
                    "employee": {
                        "id": emp.id,
                        "full_name": emp.full_name,
                        "position": emp.position,
                        "section": emp.section,
                    },
                    "work_logs": work_logs_data,
                    "summary": {
                        "total_hours": round(total_hours, 2),
                        "working_days": working_days,
                        "overtime_hours": round(overtime_hours, 2),
                        "paid_holidays": paid_holidays,
                        "unpaid_holidays": unpaid_holidays,
                    },
                }
            )

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(
                {
                    "success": True,
                    "employees": employee_logs,
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "filter_type": filter_type,
                }
            )

        return render_template(
            "work.html",
            employees=employee_logs,
            today=today,
            filter_type=filter_type,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
        )
    except Exception as e:
        Config.logger.error(f" Error in work: {e}")
        return jsonify({"error": str(e)}), 500


@worklog_bp.route("/check_in/<int:employee_id>", methods=["POST"])
@role_required(["worker", "admin"])
def check_in(employee_id):
    try:
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        work_log = WorkLog.query.filter_by(
            employee_id=employee_id, log_date=date.today()
        ).first()
        if not work_log:
            work_log = WorkLog(
                employee_id=employee_id, log_date=date.today(), worked_hours=0.0
            )
            db.session.add(work_log)
            db.session.commit()

        if work_log.check_in_time:
            return jsonify({"error": "Check-in already recorded for today"}), 400

        work_log.check_in_time = datetime.now()
        db.session.commit()
        logger.info(f"Check-in for employee {employee_id} at {work_log.check_in_time}")
        return (
            jsonify(
                {
                    "message": "Check-in recorded successfully",
                    "check_in_time": work_log.check_in_time.strftime("%H:%M"),
                }
            ),
            200,
        )
    except Exception as e:
        db.session.rollback()
        logger.error(f"Check-in error: {e}")
        return jsonify({"error": str(e)}), 500


@worklog_bp.route("/check_out/<int:employee_id>", methods=["POST"])
@role_required(["worker", "admin"])
def check_out(employee_id):
    try:
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({"error": "Employee not found"}), 404

        work_log = WorkLog.query.filter_by(
            employee_id=employee_id, log_date=date.today()
        ).first()
        if not work_log:
            return jsonify({"error": "No check-in recorded for today"}), 400
        if work_log.check_out_time:
            return jsonify({"error": "Check-out already recorded for today"}), 400

        work_log.check_out_time = datetime.now()
        if work_log.check_in_time and work_log.check_out_time:
            # Direct subtraction of datetime objects
            time_diff = work_log.check_out_time - work_log.check_in_time
            work_log.worked_hours = round(time_diff.total_seconds() / 3600, 2)
        db.session.commit()
        logger.info(
            f"Check-out for employee {employee_id} at {work_log.check_out_time}"
        )
        return (
            jsonify(
                {
                    "message": "Check-out recorded successfully",
                    "check_out_time": work_log.check_out_time.strftime("%H:%M"),
                    "daily_hours": work_log.worked_hours,
                }
            ),
            200,
        )
    except Exception as e:
        db.session.rollback()
        logger.error(f"Check-out error: {e}")
        return jsonify({"error": str(e)}), 500


@worklog_bp.route("/get_work_logs", methods=["GET"])
@role_required(["admin"])
def get_work_logs():
    try:
        logs = WorkLog.query.order_by(WorkLog.log_date).all()
        work_logs_list = [
            {
                "id": log.id,
                "employee_id": log.employee_id,
                "log_date": log.log_date.strftime("%Y-%m-%d"),
                "check_in_time": (
                    log.check_in_time.strftime("%H:%M")
                    if log.check_in_time
                    else "--:--"
                ),
                "check_out_time": (
                    log.check_out_time.strftime("%H:%M")
                    if log.check_out_time
                    else "--:--"
                ),
                "worked_hours": log.worked_hours or 0,
                "holidays": log.holidays,
            }
            for log in logs
        ]
        return jsonify({"success": True, "logs": work_logs_list})
    except Exception as e:
        Config.logger.error(f" Error in get_work_logs: {e}")
        return jsonify({"error": str(e)}), 500


@worklog_bp.route("/update_work_log/<int:log_id>", methods=["POST"])
@role_required(["admin"])
def update_work_log(log_id):
    try:
        data = request.get_json()
        holiday_status = data.get("holiday_status")
        check_in_time = data.get("check_in_time")
        check_out_time = data.get("check_out_time")
        reset_worklog = data.get("reset_worklog", False)

        work_log = WorkLog.query.get_or_404(log_id)

        if holiday_status:
            work_log.holidays = holiday_status

        if reset_worklog or holiday_status in ("paid", "unpaid", "weekend"):
            work_log.check_in_time = None
            work_log.check_out_time = None
            work_log.worked_hours = 0
        else:
            if check_in_time:
                # Преобразуем строку в time
                t_in = datetime.strptime(check_in_time, "%H:%M").time()
                work_log.check_in_time = datetime.combine(work_log.log_date, t_in)
            if check_out_time:
                # Преобразуем строку в time
                t_out = datetime.strptime(check_out_time, "%H:%M").time()
                work_log.check_out_time = datetime.combine(work_log.log_date, t_out)

            if work_log.check_in_time and work_log.check_out_time:
                delta = work_log.check_out_time - work_log.check_in_time
                work_log.worked_hours = round(delta.total_seconds() / 3600, 2)

        db.session.commit()
        summary = calculate_employee_summary(work_log.employee_id)

        return jsonify(
            {
                "success": True,
                "log_id": log_id,
                "updated_check_in": (
                    work_log.check_in_time.strftime("%H:%M")
                    if work_log.check_in_time
                    else "--:--"
                ),
                "updated_check_out": (
                    work_log.check_out_time.strftime("%H:%M")
                    if work_log.check_out_time
                    else "--:--"
                ),
                "updated_worked_hours": f"{work_log.worked_hours:.2f}",
                "summary": {
                    "total_hours": summary["total_hours"],
                    "working_days": summary["working_days"],
                    "overtime_hours": summary["overtime_hours"],
                    "paid_holidays": summary["paid_holidays"],
                    "unpaid_holidays": summary["unpaid_holidays"],
                },
            }
        )
    except Exception as e:
        Config.logger.error(f" Error in update_work_log: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


def add_work_log(employee_id):
    try:
        data = request.get_json()
        log_date = data.get("log_date")
        check_in_time = data.get("check_in_time")
        check_out_time = data.get("check_out_time")

        if not all([log_date, check_in_time, check_out_time]):
            return (
                jsonify({"success": False, "message": "Missing required fields"}),
                400,
            )

        # Преобразуем дату
        log_date = datetime.strptime(log_date, "%Y-%m-%d").date()

        try:
            # Преобразуем время в объект времени
            check_in_time = datetime.strptime(check_in_time, "%H:%M").time()
            check_out_time = datetime.strptime(check_out_time, "%H:%M").time()
        except ValueError:
            return (
                jsonify(
                    {"success": False, "message": "Invalid time format, should be HH:MM"}
                ),
                400,
            )

        # Создаем datetime объект, комбинируя дату и время
        check_in_datetime = datetime.combine(log_date, check_in_time)
        check_out_datetime = datetime.combine(log_date, check_out_time)

        # Проверяем наличие уже существующего рабочего дня
        existing_log = WorkLog.query.filter_by(
            employee_id=employee_id, log_date=log_date
        ).first()
        if existing_log:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Work log for this date already exists",
                    }
                ),
                400,
            )

        # Создаем новый рабочий лог
        new_log = WorkLog(
            employee_id=employee_id,
            log_date=log_date,
            check_in_time=check_in_datetime,
            check_out_time=check_out_datetime,
            holidays="workingday",
        )
        new_log.calculate_worked_hours()
        db.session.add(new_log)
        db.session.commit()

        return jsonify({"success": True, "message": "Work log added successfully"})
    except ValueError as e:
        Config.logger.error(f" Error parsing date/time: {e}")
        db.session.rollback()
        return (
            jsonify({"success": False, "message": "Invalid date or time format"}),
            400,
        )
    except Exception as e:
        Config.logger.error(f" Error adding work log: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@worklog_bp.route("/export_excel", methods=["POST"])
@role_required(["admin"])
def export_excel():
    try:
        data = request.get_json()
        selected_employees = data.get("selectedEmployees", [])
        group = data.get("group")
        filter_type = data.get("filterType")
        start_date_str = data.get("startDate")
        end_date_str = data.get("endDate")
        status_filter = data.get("statusFilter")

        if not selected_employees:
            return jsonify({"message": "No employees selected for export"}), 400

        employees_query = Employee.query.filter(Employee.id.in_(selected_employees))
        if group:
            employees_query = employees_query.filter(Employee.section == group)
        employees = employees_query.all()

        today = date.today()

        if filter_type == "today":
            start_date = end_date = today
        elif filter_type == "yesterday":
            start_date = end_date = today - timedelta(days=1)
        elif filter_type == "last7days":
            start_date = today - timedelta(days=6)
            end_date = today
        elif filter_type == "last30days":
            start_date = today - timedelta(days=29)
            end_date = today
        elif filter_type == "thismonth":
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(
                days=1
            )
        elif filter_type == "lastmonth":
            first_day_this = today.replace(day=1)
            end_date = first_day_this - timedelta(days=1)
            start_date = end_date.replace(day=1)
        elif (
            filter_type in ("custom", "personalizado")
            and start_date_str
            and end_date_str
        ):
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                if start_date > end_date:
                    raise ValueError("startDate > endDate")
            except ValueError:
                return jsonify({"message": "Invalid custom date range"}), 400
        else:
            start_date = end_date = today

        employees_data = []
        for emp in employees:
            q = WorkLog.query.filter(
                WorkLog.employee_id == emp.id,
                WorkLog.log_date >= start_date,
                WorkLog.log_date <= end_date,
            )
            if status_filter:
                q = q.filter(WorkLog.holidays == status_filter)
            logs = q.order_by(WorkLog.log_date).all()
            employees_data.append({"employee": emp, "logs": logs})

        wb = Workbook()
        ws = wb.active
        ws.title = "Work Logs"

        bold = Font(bold=True)
        center = Alignment(horizontal="center")

        row = 1
        for block in employees_data:
            emp = block["employee"]
            logs = block["logs"]

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(
                row=row, column=1, value=f"{emp.full_name} - {emp.position} ({emp.nie})"
            ).font = bold
            row += 1

            headers = [
                "Fecha",
                "Entrada",
                "Salida",
                "Total",
                "Extra",
                "Tipo Día",
                "Días trabajados",
            ]
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = bold
                c.alignment = center
            row += 1

            total_days = total_hours = total_overtime = 0
            for log in logs:
                ws.cell(row=row, column=1, value=log.log_date.strftime("%Y-%m-%d"))
                ws.cell(
                    row=row,
                    column=2,
                    value=(
                        log.check_in_time.strftime("%H:%M")
                        if log.check_in_time
                        else "--:--"
                    ),
                )
                ws.cell(
                    row=row,
                    column=3,
                    value=(
                        log.check_out_time.strftime("%H:%M")
                        if log.check_out_time
                        else "--:--"
                    ),
                )
                hours = round(log.worked_hours or 0, 2)
                overtime = max(0, round(hours - 8, 2))
                ws.cell(row=row, column=4, value=hours)
                ws.cell(row=row, column=5, value=overtime)
                ws.cell(
                    row=row, column=6, value=(log.holidays or "workingday").capitalize()
                )
                is_work = bool(log.check_in_time and log.check_out_time)
                ws.cell(row=row, column=7, value=1 if is_work else 0)

                if is_work:
                    total_days += 1
                    total_hours += hours
                    total_overtime += overtime

                row += 1

            for label, val in [
                ("Días totales", total_days),
                ("Horas totales", total_hours),
                ("Horas extra", total_overtime),
            ]:
                ws.cell(row=row, column=5, value=label).font = bold
                ws.cell(row=row, column=6, value=val)
                row += 1

            row += 1

        for col in ws.columns:
            length = max(
                (len(str(cell.value)) for cell in col if cell.value), default=0
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

        out_path = os.path.join("instance", "filtered_employee_logs.xlsx")
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)

        return send_file(
            out_path, as_attachment=True, download_name="filtered_employee_logs.xlsx"
        )
    except Exception as e:
        Config.logger.error(f" Error in export_excel: {e}")
        return jsonify({"error": str(e)}), 500
