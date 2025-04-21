from flask import Flask, render_template, session, redirect, url_for, jsonify, request
from functools import wraps
from typing import Union, Tuple
import pandas as pd
import io
from flask import send_file
from flask_sqlalchemy import SQLAlchemy
from flask_apscheduler import APScheduler
from datetime import datetime, date, timedelta, time  # Ensure 'time' is imported
import logging
import time  # For timing measurements
import sys
from models import Employee, Admin, DashboardUser, WorkLog, db
from config import Config
from sqlalchemy import and_
# start_time = time.time()
print("🔄 Запуск приложения...")
import logging
import sys
# test
# Настраиваем логирование
logging.basicConfig(
    level=logging.DEBUG,  # DEBUG показывает всё, INFO может скрывать ошибки
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]  # Логи в консоль
)
logger = logging.getLogger(__name__)

logger.info("✅ Логирование настроено! Это тестовый лог.")
# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# mid_time = time.time()
# print(f"Импорт модулей занял: {mid_time - start_time:.2f} сек.")
# app_init = time.time()
# print(f"🚀 Инициализация Flask заняла: {app_init - mid_imports:.2f} сек.")

# Проверка конфигурации
if not app.config.get('SQLALCHEMY_DATABASE_URI'):
    logger.error("❌ SQLALCHEMY_DATABASE_URI не настроен в config.py")
    sys.exit(1)

# Инициализация базы данных
# start_time = time.time()
logger.info(f"✅ Подключение к базе: {app.config['SQLALCHEMY_DATABASE_URI']}")
try:
    db.init_app(app)
except Exception as e:
    logger.error(f"❌ Ошибка инициализации базы данных: {e}")
    sys.exit(1)
# logger.info(f"⏱️ Инициализация базы данных заняла: {time.time() - start_time:.2f} секунд")

# Инициализация APScheduler
start_time = time.time()
scheduler = APScheduler()
try:
    scheduler.init_app(app)
except Exception as e:
    logger.error(f"❌ Ошибка инициализации планировщика: {e}")
    sys.exit(1)
# logger.info(f"⏱️ Инициализация планировщика заняла: {time.time() - start_time:.2f} секунд")

# Создание таблиц при первом запуске
start_time = time.time()
with app.app_context():
    try:
        db.create_all()
        logger.info("✅ Таблицы базы данных созданы")
    except Exception as e:
        logger.error(f"❌ Ошибка создания таблиц: {e}")
        sys.exit(1)
logger.info(f"⏱️ Создание таблиц заняло: {time.time() - start_time:.2f} секунд")

# Комбинированная функция для сброса и создания логов
def reset_and_create_logs():
    with app.app_context():
        try:
            start_time = time.time()
            today = date.today()
            employees = Employee.query.all()
            for emp in employees:
                work_log = WorkLog.query.filter_by(employee_id=emp.id, log_date=today).first()
                if work_log:
                    work_log.check_in_time = None
                    work_log.check_out_time = None
                    work_log.worked_hours = 0
                else:
                    new_log = WorkLog(employee_id=emp.id, log_date=today, worked_hours=0)
                    db.session.add(new_log)
            db.session.commit()
            logger.info(f"🔄 Данные сотрудников сброшены и созданы логи для {today}")
            logger.info(f"⏱️ reset_and_create_logs занял: {time.time() - start_time:.2f} секунд")
        except Exception as e:
            logger.error(f"❌ Ошибка в reset_and_create_logs: {e}")
            db.session.rollback()

# Запуск планировщика
start_time = time.time()
if not scheduler.running:
    try:
        scheduler.add_job(
            id="reset_and_create_logs",
            func=reset_and_create_logs,
            trigger="cron",
            hour=0, minute=1
        )
        scheduler.start()
        logger.info("✅ Планировщик запущен")
    except Exception as e:
        logger.error(f"❌ Ошибка запуска планировщика: {e}")
        sys.exit(1)
logger.info(f"⏱️ Запуск планировщика занял: {time.time() - start_time:.2f} секунд")

# Декоратор проверки ролей
def role_required(allowed_roles):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user_role = session.get('role')
            if not user_role:
                return redirect(url_for('worker_login' if 'worker' in allowed_roles else 'admin_login'))
            if user_role not in allowed_roles:
                return redirect(url_for('board' if user_role == 'worker' else 'admin_login'))
            return func(*args, **kwargs)
        return wrapper
    return decorator

# Роуты
@app.route('/')
def home():
    return render_template('home.html')

@app.route('/dashboard')
@role_required(['admin'])
def admin_panel():
    try:
        employees = Employee.query.all()
        logger.info(f"🔍 Загруженные сотрудники: {len(employees)}")
        return render_template('index.html', employees=employees)
    except Exception as e:
        logger.error(f"❌ Ошибка в admin_panel: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/board')
@role_required(['worker', 'admin'])
def board():
    try:
        today = date.today()
        employees = Employee.query.all()
        dashboard_data = []

        for employee in employees:
            work_log = WorkLog.query.filter_by(employee_id=employee.id, log_date=today).first()
            if not work_log:
                work_log = WorkLog(employee_id=employee.id, log_date=today, worked_hours=0.0)
                db.session.add(work_log)
                db.session.commit()

            dashboard_data.append({
                'employee': employee,
                'check_in_time': work_log.check_in_time.strftime('%H:%M') if work_log.check_in_time else '--:--',
                'check_out_time': work_log.check_out_time.strftime('%H:%M') if work_log.check_out_time else '--:--',
                'daily_hours': round(work_log.worked_hours or 0, 2)
            })

        return render_template('board.html', dashboard_data=dashboard_data, current_date=today)
    except Exception as e:
        logger.error(f"❌ Ошибка в board: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/dashboard_data')
def get_dashboard_data():
    try:
        today = date.today()
        employees = Employee.query.all()
        data = []

        for emp in employees:
            work_log = WorkLog.query.filter_by(employee_id=emp.id, log_date=today).first()
            if not work_log:
                work_log = WorkLog(employee_id=emp.id, log_date=today, worked_hours=0)
                db.session.add(work_log)
                db.session.commit()

            data.append({
                "id": emp.id,
                "full_name": emp.full_name,
                "check_in_time": work_log.check_in_time.strftime("%H:%M") if work_log.check_in_time else "--:--",
                "check_out_time": work_log.check_out_time.strftime("%H:%M") if work_log.check_out_time else "--:--",
                "daily_hours": round(work_log.worked_hours or 0, 2)
            })

        return jsonify(data)
    except Exception as e:
        logger.error(f"❌ Ошибка в get_dashboard_data: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/admin_register', methods=['GET', 'POST'])
def admin_register():
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            password = request.form.get('password')
            confirm_password = request.form.get('confirm_password')

            if password != confirm_password:
                return jsonify({'error': 'Passwords do not match'}), 400

            if Admin.query.filter_by(email=email).first():
                return jsonify({'error': 'Admin with this email already exists'}), 400

            admin = Admin(email=email)
            admin.set_password(password)
            db.session.add(admin)
            db.session.commit()
            return redirect(url_for('admin_login'))

        return render_template('admin_register.html')
    except Exception as e:
        logger.error(f"❌ Ошибка в admin_register: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    try:
        if request.method == 'POST':
            email = request.form.get('email')
            password = request.form.get('password')
            admin = Admin.query.filter_by(email=email).first()

            if admin and admin.check_password(password):
                session['role'] = 'admin'
                session['admin_email'] = email
                return redirect(url_for('admin_panel'))
            return jsonify({'error': 'Invalid credentials'}), 401

        return render_template('admin_login.html')
    except Exception as e:
        logger.error(f"❌ Ошибка в admin_login: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/worker_register', methods=['GET', 'POST'])
def worker_register():
    try:
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            confirm_password = request.form.get('confirm_password')

            if password != confirm_password:
                return jsonify({'error': 'Passwords do not match'}), 400

            if DashboardUser.query.filter_by(username=username).first():
                return jsonify({'error': 'Worker with this username already exists'}), 400

            worker = DashboardUser(username=username, role='worker')
            worker.set_password(password)
            db.session.add(worker)
            db.session.commit()
            return redirect(url_for('worker_login'))

        return render_template('worker_register.html')
    except Exception as e:
        logger.error(f"❌ Ошибка в worker_register: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/worker_login', methods=['GET', 'POST'])
def worker_login():
    try:
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            user = DashboardUser.query.filter_by(username=username).first()

            if user and user.check_password(password):
                session.permanent = True
                session['user_id'] = user.id
                session['role'] = user.role
                return redirect(url_for('board' if user.role == 'worker' else 'admin_panel'))
            return jsonify({'error': 'Invalid credentials'}), 401

        return render_template('worker_login.html')
    except Exception as e:
        logger.error(f"❌ Ошибка в worker_login: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/logout')
def logout():
    try:
        session.clear()
        return redirect(url_for('admin_login'))
    except Exception as e:
        logger.error(f"❌ Ошибка в logout: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/worker_logout')
def worker_logout():
    try:
        session.clear()
        return redirect(url_for('worker_login'))
    except Exception as e:
        logger.error(f"❌ Ошибка в worker_logout: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/employees')
def get_employees():
    try:
        employees = Employee.query.all()
        return jsonify([{
            "id": emp.id,
            "full_name": emp.full_name,
            "nie": emp.nie,
            "position": emp.position,
            "phone": emp.phone,
            "email": emp.email,
            "start_date": emp.start_date.strftime("%Y-%m-%d") if emp.start_date else "",
            "work_start_time": emp.work_start_time.strftime("%H:%M") if emp.work_start_time else "",
            "work_end_time": emp.work_end_time.strftime("%H:%M") if emp.work_end_time else ""
        } for emp in employees])
    except Exception as e:
        logger.error(f"❌ Ошибка в get_employees: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/add', methods=['POST'])
def add_employee():
    try:
        nie = request.form.get('nie')
        email = request.form.get('email')

        if Employee.query.filter((Employee.nie == nie) | (Employee.email == email)).first():
            return jsonify({'error': 'Employee with this NIE or Email already exists'}), 400

        full_name = request.form.get('full_name')
        position = request.form.get('position')
        phone = request.form.get('phone')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        section = request.form.get('section')
        work_start_time = request.form.get('work_start_time')
        work_end_time = request.form.get('work_end_time')

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
            work_start_time = datetime.strptime(work_start_time, '%H:%M').time() if work_start_time else None
            work_end_time = datetime.strptime(work_end_time, '%H:%M').time() if work_end_time else None
        except ValueError:
            return jsonify({'error': 'Invalid date or time format'}), 400

        new_employee = Employee(
            full_name=full_name, nie=nie, position=position, phone=phone, email=email,
            start_date=start_date, end_date=end_date, section=section,
            work_start_time=work_start_time, work_end_time=work_end_time
        )
        db.session.add(new_employee)
        db.session.commit()
        return jsonify({'message': 'Employee added successfully!', 'employee': {
            "id": new_employee.id, "full_name": new_employee.full_name, "nie": new_employee.nie,
            "position": new_employee.position, "phone": new_employee.phone, "email": new_employee.email,
            "start_date": new_employee.start_date.strftime("%Y-%m-%d") if new_employee.start_date else "",
            "work_start_time": new_employee.work_start_time.strftime("%H:%M") if new_employee.work_start_time else "",
            "work_end_time": new_employee.work_end_time.strftime("%H:%M") if new_employee.work_end_time else ""
        }})
    except Exception as e:
        logger.error(f"❌ Ошибка в add_employee: {e}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/edit/<int:id>', methods=['POST'])
def edit_employee(id):
    try:
        employee = db.session.get(Employee, id)
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404

        try:
            employee.full_name = request.form.get('full_name', employee.full_name)
            employee.nie = request.form.get('nie', employee.nie)
            employee.phone = request.form.get('phone', employee.phone)
            employee.position = request.form.get('position', employee.position)
            employee.email = request.form.get('email', employee.email)
            employee.section = request.form.get('section', employee.section)
            employee.days_per_week = int(request.form.get('days_per_week', employee.days_per_week or 0))

            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            work_start_time_str = request.form.get('work_start_time')
            work_end_time_str = request.form.get('work_end_time')

            employee.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else employee.start_date
            employee.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else employee.end_date
            employee.work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time() if work_start_time_str else employee.work_start_time
            employee.work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time() if work_end_time_str else employee.work_end_time

            db.session.commit()
            logger.info(f"Employee {employee.full_name} updated successfully")
            return jsonify({'message': 'Employee updated successfully'}), 200
        except ValueError as e:
            logger.error(f"Error parsing data: {e}")
            return jsonify({'error': 'Invalid data format'}), 400
    except Exception as e:
        logger.error(f"Error updating employee: {e}")
        db.session.rollback()
        return jsonify({'error': 'Server error'}), 500

@app.route('/employee/<int:employee_id>', methods=['GET'])
def get_employee(employee_id):
    try:
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404

        return jsonify({
            'id': employee.id, 'full_name': employee.full_name, 'nie': employee.nie,
            'start_date': employee.start_date.strftime('%Y-%m-%d') if employee.start_date else '',
            'end_date': employee.end_date.strftime('%Y-%m-%d') if employee.end_date else '',
            'work_start_time': employee.work_start_time.strftime('%H:%M') if employee.work_start_time else '',
            'work_end_time': employee.work_end_time.strftime('%H:%M') if employee.work_end_time else '',
            'days_per_week': employee.days_per_week or 0, 'position': employee.position,
            'section': employee.section, 'phone': employee.phone, 'email': employee.email
        })
    except Exception as e:
        logger.error(f"❌ Ошибка в get_employee: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/check_in/<int:employee_id>', methods=['POST'])
def check_in(employee_id):
    try:
        current_date = date.today()
        work_log = WorkLog.query.filter_by(employee_id=employee_id, log_date=current_date).first()

        if work_log and work_log.check_in_time:
            return jsonify({"error": "Already checked in for today"}), 400

        if not work_log:
            work_log = WorkLog(employee_id=employee_id, log_date=current_date)

        work_log.check_in_time = datetime.now()
        db.session.add(work_log)
        db.session.commit()
        logger.info(f"Check-in for employee {employee_id} at {work_log.check_in_time}")
        return jsonify({"check_in_time": work_log.check_in_time.strftime('%H:%M')}), 200
    except Exception as e:
        logger.error(f"Check-in error: {e}")
        db.session.rollback()
        return jsonify({"error": "Server error"}), 500

@app.route('/check_out/<int:employee_id>', methods=['POST'])
def check_out(employee_id):
    try:
        current_date = date.today()
        work_log = WorkLog.query.filter_by(employee_id=employee_id, log_date=current_date).first()

        if not work_log or not work_log.check_in_time:
            return jsonify({"error": "Cannot check out before checking in"}), 400
        if work_log.check_out_time:
            return jsonify({"error": "Already checked out for today"}), 400

        work_log.check_out_time = datetime.now()
        work_log.calculate_worked_hours()
        db.session.commit()
        logger.info(f"Check-out for employee {employee_id} at {work_log.check_out_time}")
        return jsonify({
            "check_out_time": work_log.check_out_time.strftime('%H:%M'),
            "daily_hours": round(work_log.worked_hours, 2)
        }), 200
    except Exception as e:
        logger.error(f"Check-out error: {e}")
        db.session.rollback()
        return jsonify({"error": "Server error"}), 500

@app.route('/delete/<int:employee_id>', methods=['POST'])
def delete_employee(employee_id):
    try:
        employee = db.session.get(Employee, employee_id)
        if not employee:
            return jsonify({'error': 'Employee not found'}), 404
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'message': 'Employee deleted successfully'}), 200
    except Exception as e:
        logger.error(f"Delete error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Server error'}), 500

from flask import request, jsonify, session, render_template
from datetime import datetime, date, timedelta
from sqlalchemy import and_
import calendar

@app.route('/work', methods=['GET'])
def work():
    try:
        # 🛠 Обработка filter с учётом personalizado и сброса сессии
        requested = request.args.get('filter')
        allowed = {'today', 'yesterday', 'last7days', 'last30days', 'thismonth', 'lastmonth'}

        if requested in allowed:
            filter_type = requested
            session['filter_type'] = filter_type
        elif requested == 'personalizado':
            filter_type = 'personalizado'
            # очищаем предыдущий обычный фильтр, чтобы он не применялся по умолчанию
            session.pop('filter_type', None)
        else:
            prev = session.get('filter_type')
            if prev in allowed:
                filter_type = prev
            else:
                filter_type = 'thismonth'
                session['filter_type'] = filter_type

        today = date.today()
        start_date = end_date = None

        # Основные фильтры
        if filter_type == 'today':
            start_date = end_date = today
        elif filter_type == 'yesterday':
            start_date = end_date = today - timedelta(days=1)
        elif filter_type == 'last7days':
            start_date = today - timedelta(days=6)
            end_date = today
        elif filter_type == 'last30days':
            start_date = today - timedelta(days=29)
            end_date = today
        elif filter_type == 'thismonth':
            start_date = today.replace(day=1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = today.replace(day=last_day)
        elif filter_type == 'lastmonth':
            first_day_this = today.replace(day=1)
            last_day_prev = first_day_this - timedelta(days=1)
            start_date = last_day_prev.replace(day=1)
            end_date = last_day_prev
        elif filter_type == 'personalizado':
            # кастомный диапазон при нажатии "Давай"
            start_str = request.args.get('start_date')
            end_str   = request.args.get('end_date')

            if not start_str or not end_str:
                return jsonify({"error": "El filtro 'personalizado' requiere 'start_date' y 'end_date'"}), 400

            try:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date   = datetime.strptime(end_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({"error": "Formato de fecha no válido"}), 400

            if start_date > end_date:
                return jsonify({"error": "La fecha de inicio no puede ser después de la fecha de finalización"}), 400
        else:
            # fallback — сбрасываем на thismonth
            filter_type = 'thismonth'
            start_date = today.replace(day=1)
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = today.replace(day=last_day)
            session['filter_type'] = filter_type

        # Проверка наличия дат
        if start_date is None or end_date is None:
            return jsonify({"error": "Faltan las fechas de filtro"}), 400

        # Загрузка сотрудников и логов
        employees = Employee.query.all()
        employee_logs = []

        for emp in employees:
            logs = WorkLog.query.filter(
                and_(
                    WorkLog.employee_id == emp.id,
                    WorkLog.log_date >= start_date,
                    WorkLog.log_date <= end_date
                )
            ).all()

            if not logs:
                continue

            total_hours = working_days = paid_holidays = unpaid_holidays = overtime_hours = 0
            work_logs_data = []

            for log in logs:
                log.calculate_worked_hours()
                wd = log.log_date
                if isinstance(wd, str):
                    wd = datetime.strptime(wd, '%Y-%m-%d').date()

                hours = float(log.worked_hours or 0)
                total_hours += hours

                if log.holidays == 'workingday' and hours > 0:
                    working_days += 1
                    if hours > 8:
                        overtime_hours += (hours - 8)
                elif log.holidays == 'paid':
                    paid_holidays += 1
                elif log.holidays == 'unpaid':
                    unpaid_holidays += 1

                work_logs_data.append({
                    'id': log.id,
                    'log_date': wd.strftime('%Y-%m-%d'),
                    'check_in_time': log.check_in_time.strftime('%H:%M') if log.check_in_time else '--:--',
                    'check_out_time': log.check_out_time.strftime('%H:%M') if log.check_out_time else '--:--',
                    'worked_hours': round(hours, 2),
                    'holidays': log.holidays or 'workingday'
                })

            employee_logs.append({
                'employee': {
                    'id': emp.id,
                    'full_name': emp.full_name,
                    'position': emp.position,
                    'section': emp.section
                },
                'work_logs': work_logs_data,
                'summary': {
                    'total_hours': round(total_hours, 2),
                    'working_days': working_days,
                    'overtime_hours': round(overtime_hours, 2),
                    'paid_holidays': paid_holidays,
                    'unpaid_holidays': unpaid_holidays
                }
            })

        # Ответ для AJAX или рендер страницы
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({
                'success': True,
                'employees': employee_logs,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'filter_type': filter_type
            })

        return render_template(
            'work.html',
            employees=employee_logs,
            today=today,
            filter_type=filter_type,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat()
        )

    except Exception as e:
        return jsonify({"error": f"❌ Ошибка в /work: {str(e)}"}), 500
# iuggigi
def calculate_employee_summary(employee_id):
    """
    Пересчитывает суммарные часы работы, дни, отпуска для сотрудника.
    """
    work_logs = WorkLog.query.filter_by(employee_id=employee_id).all()

    total_hours = sum(log.worked_hours or 0 for log in work_logs)
    working_days = sum(1 for log in work_logs if log.worked_hours and log.worked_hours > 0)
    paid_holidays = sum(1 for log in work_logs if log.holidays == 'paid')
    unpaid_holidays = sum(1 for log in work_logs if log.holidays == 'unpaid')

    # ✅ Часы округляем в сторону большего значения
    total_hours = round(total_hours)
    overtime_hours = max(0, total_hours - (working_days * 8))

    logger.info(f"📌 Пересчёт саммари для employee_id={employee_id}:")
    logger.info(f"✅ total_hours={total_hours}, working_days={working_days}, overtime_hours={overtime_hours}")
    logger.info(f"✅ paid_holidays={paid_holidays}, unpaid_holidays={unpaid_holidays}")

    return {
        "total_hours": total_hours,
        "working_days": working_days,
        "overtime_hours": overtime_hours,
        "paid_holidays": paid_holidays,
        "unpaid_holidays": unpaid_holidays
    }


@app.route('/get_work_logs', methods=['GET'])
def get_work_logs():
    """
    Возвращает все ворклоги в JSON-формате, отсортированные по дате.
    """
    logs = WorkLog.query.order_by(WorkLog.log_date).all()

    work_logs_list = [
        {
            "id": log.id,
            "employee_id": log.employee_id,
            "log_date": log.log_date.strftime('%Y-%m-%d'),
            "check_in_time": log.check_in_time.strftime('%H:%M') if log.check_in_time else "--:--",
            "check_out_time": log.check_out_time.strftime('%H:%M') if log.check_out_time else "--:--",
            "worked_hours": log.worked_hours or 0,
            "holidays": log.holidays
        }
        for log in logs
    ]

    return jsonify({"success": True, "logs": work_logs_list})
# в начале файла добавить
from datetime import datetime
# hjji
@app.route('/update_work_log/<int:log_id>', methods=['POST'])
def update_work_log(log_id):
    data = request.get_json()
    holiday_status = data.get("holiday_status")
    check_in_time  = data.get("check_in_time")   # строка "HH:MM" или None
    check_out_time = data.get("check_out_time")  # строка "HH:MM" или None
    reset_worklog  = data.get("reset_worklog", False)

    work_log = WorkLog.query.get_or_404(log_id)

    # обновляем статус дня
    if holiday_status:
        work_log.holidays = holiday_status

    if reset_worklog:
        # сбрасываем время и часы
        work_log.check_in_time  = None
        work_log.check_out_time = None
        work_log.worked_hours   = 0
    else:
        # если пришло новое время — конвертим в datetime с датой log_date
        if check_in_time:
            t_in = datetime.strptime(check_in_time, "%H:%M").time()
            work_log.check_in_time = datetime.combine(work_log.log_date, t_in)
        if check_out_time:
            t_out = datetime.strptime(check_out_time, "%H:%M").time()
            work_log.check_out_time = datetime.combine(work_log.log_date, t_out)

        # пересчёт worked_hours
        if work_log.check_in_time and work_log.check_out_time:
            delta = work_log.check_out_time - work_log.check_in_time
            work_log.worked_hours = round(delta.total_seconds() / 3600, 2)

    db.session.commit()

    return jsonify({
        "success": True,
        "log_id": log_id,
        "updated_check_in": work_log.check_in_time.strftime('%H:%M') if work_log.check_in_time else "--:--",
        "updated_check_out": work_log.check_out_time.strftime('%H:%M') if work_log.check_out_time else "--:--",
        "updated_worked_hours": f"{work_log.worked_hours:.2f}"
    })

# Функция для определения диапазона дат по фильтру
def get_date_range(filter_type: str) -> Tuple[date, date]:
    today = date.today()

    if filter_type == 'today':
        return today, today
    elif filter_type == 'yesterday':
        return today - timedelta(days=1), today - timedelta(days=1)
    elif filter_type == 'last7days':
        return today - timedelta(days=6), today
    elif filter_type == 'last30days':
        return today - timedelta(days=29), today
    elif filter_type == 'thismonth':
        start_date = today.replace(day=1)
        end_date = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        return start_date, end_date
    elif filter_type == 'lastmonth':
        first_day = today.replace(day=1)
        end_date = first_day - timedelta(days=1)
        start_date = end_date.replace(day=1)
        return start_date, end_date
    return today, today  # Если фильтр не распознан, возвращаем сегодня

# Фильтр для форматирования часов
@app.template_filter('format_hours')
def format_hours(value):
    if value is None or value == 0:
        return '0h 0min'
    hours = int(value)
    minutes = int((value - hours) * 60)
    return f'{hours}h {minutes}min'

@app.route('/add_work_log/<int:employee_id>', methods=['POST'])
def add_work_log(employee_id):
    try:
        data = request.get_json()
        log_date = data.get('log_date')
        check_in_time = data.get('check_in_time')
        check_out_time = data.get('check_out_time')

        if not all([log_date, check_in_time, check_out_time]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400

        # Преобразуем даты и время
        log_date = datetime.strptime(log_date, '%Y-%m-%d').date()
        check_in_time = datetime.strptime(check_in_time, '%H:%M').time()
        check_out_time = datetime.strptime(check_out_time, '%H:%M').time()

        # Combine log_date with check_in_time and check_out_time to create full datetime objects
        check_in_datetime = datetime.combine(log_date, check_in_time)
        check_out_datetime = datetime.combine(log_date, check_out_time)

        # Проверяем, существует ли уже запись для этой даты
        existing_log = WorkLog.query.filter_by(employee_id=employee_id, log_date=log_date).first()
        if existing_log:
            return jsonify({'success': False, 'message': 'Work log for this date already exists'}), 400

        # Создаем новый WorkLog
        new_log = WorkLog(
            employee_id=employee_id,
            log_date=log_date,
            check_in_time=check_in_datetime,  # Store as TIMESTAMP
            check_out_time=check_out_datetime,  # Store as TIMESTAMP
            holidays='workingday'  # Значение по умолчанию
        )
        new_log.calculate_worked_hours()  # Рассчитываем часы работы
        db.session.add(new_log)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Work log added successfully'})

    except ValueError as e:
        logger.error(f"❌ Ошибка при парсинге даты/времени: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Invalid date or time format'}), 400
    except Exception as e:
        logger.error(f"❌ Ошибка при добавлении work log: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

import os
from flask import send_file, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date

from datetime import datetime, date, timedelta
import os
from flask import request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

@app.route('/export_excel', methods=['POST'])
def export_excel():
    """
    Экспорт ворклогов сотрудников в Excel с учётом фильтров.
    """
    data = request.get_json()
    selected_employees = data.get('selectedEmployees', [])
    group              = data.get('group')
    filter_type        = data.get('filterType')
    start_date_str     = data.get('startDate')
    end_date_str       = data.get('endDate')
    status_filter      = data.get('statusFilter')

    if not selected_employees:
        return jsonify({'message': 'No employees selected for export'}), 400

    # собираем сотрудников
    employees_query = Employee.query.filter(Employee.id.in_(selected_employees))
    if group:
        employees_query = employees_query.filter(Employee.section == group)
    employees = employees_query.all()

    today = date.today()

    # 1) Определяем диапазон дат по filter_type
    if filter_type == 'today':
        start_date = end_date = today
    elif filter_type == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
    elif filter_type == 'last7days':
        start_date = today - timedelta(days=6)
        end_date   = today
    elif filter_type == 'last30days':
        start_date = today - timedelta(days=29)
        end_date   = today
    elif filter_type == 'thismonth':
        start_date = today.replace(day=1)
        end_date   = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    elif filter_type == 'lastmonth':
        first_day_this = today.replace(day=1)
        end_date       = first_day_this - timedelta(days=1)
        start_date     = end_date.replace(day=1)
    # ловим оба ключа кастомного диапазона
    elif filter_type in ('custom', 'personalizado') and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()
            if start_date > end_date:
                raise ValueError("startDate > endDate")
        except ValueError:
            return jsonify({'message': 'Invalid custom date range'}), 400
    else:
        # fallback на today, но можно кинуть ошибку или взять предыдущий фильтр
        start_date = end_date = today

    # 2) Собираем логи за этот диапазон
    employees_data = []
    for emp in employees:
        q = WorkLog.query.filter(
            WorkLog.employee_id == emp.id,
            WorkLog.log_date >= start_date,
            WorkLog.log_date <= end_date
        )
        if status_filter:
            q = q.filter(WorkLog.holidays == status_filter)
        logs = q.order_by(WorkLog.log_date).all()
        employees_data.append({'employee': emp, 'logs': logs})

    # 3) Формируем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Logs"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    row = 1
    for block in employees_data:
        emp = block['employee']
        logs = block['logs']

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value=f"{emp.full_name} - {emp.position} ({emp.nie})").font = bold
        row += 1

        headers = ["Fecha", "Entrada", "Salida", "Total", "Extra", "Tipo Día", "Días trabajados"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.alignment = center
        row += 1

        total_days = total_hours = total_overtime = 0
        for log in logs:
            ws.cell(row=row, column=1, value=log.log_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=log.check_in_time.strftime('%H:%M') if log.check_in_time else '--:--')
            ws.cell(row=row, column=3, value=log.check_out_time.strftime('%H:%M') if log.check_out_time else '--:--')
            hours = round(log.worked_hours or 0, 2)
            overtime = max(0, round(hours - 8, 2))
            ws.cell(row=row, column=4, value=hours)
            ws.cell(row=row, column=5, value=overtime)
            ws.cell(row=row, column=6, value=(log.holidays or 'workingday').capitalize())
            is_work = bool(log.check_in_time and log.check_out_time)
            ws.cell(row=row, column=7, value=1 if is_work else 0)

            if is_work:
                total_days += 1
                total_hours  += hours
                total_overtime += overtime

            row += 1

        # Итоги
        for label, val in [
            ("Días totales", total_days),
            ("Horas totales", total_hours),
            ("Horas extra", total_overtime)
        ]:
            ws.cell(row=row, column=5, value=label).font = bold
            ws.cell(row=row, column=6, value=val)
            row += 1

        row += 1  # пустая строка

    # авто‑ширина
    for col in ws.columns:
        length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    # сохраняем и отдаём
    out_path = os.path.join('instance', 'filtered_employee_logs.xlsx')
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

    return send_file(
        out_path,
        as_attachment=True,
        download_name="filtered_employee_logs.xlsx"
    )

# test
if __name__ == '__main__':
    try:
        app.run(debug=True, port=5001, host='0.0.0.0')
        logger.info("✅ Приложение запущено на порту 5005")
    except Exception as e:
        logger.error(f"❌ Ошибка запуска приложения: {e}")
        sys.exit(1)